# =========================
# 이루리 영어학원 방학 모의고사 성적분석 (Streamlit) - Local
# Sheets:
#  - students (student_id, name, grade, role)
#  - wrong_answer (응시순서, 출제기관, 회차, 응시자, 원점수, 등급, 오답)
# Optional:
#  - ebsi_stats (출제기관, 회차, 학년, 문항번호, 전국오답률)
#  - grammar_info (출제기관, 회차, 학년, 문항번호, 정답개념)
#  - 3grade (이름, 모의고사응시횟수, 듣기영역(1~17번), 독해영역(18~45번), 등급평균)
#  - admin_solution (name, solution)
# =========================

import re
import sqlite3
import hashlib
from datetime import datetime, timedelta, timezone

import pandas as pd
import streamlit as st
import plotly.express as px

# 클릭 이벤트(선택 기능) - 환경에 따라 동작이 불안정할 수 있어 fallback 처리함
try:
    from streamlit_plotly_events import plotly_events
    HAS_PLOTLY_EVENTS = True
except Exception:
    HAS_PLOTLY_EVENTS = False

# 엑셀 저장(관리자 입력)
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False


# -------------------------
# Page config
# -------------------------
st.set_page_config(page_title="이루리 영어학원 성적분석", layout="wide")

# -------------------------
# Paths & limits
# -------------------------
DB_PATH = "data/auth.db"
EXCEL_PATH = "data/iruri.xlsx"

PAIR_FAIL_LIMIT = 5
PAIR_LOCK_MIN = 10
IP_FAIL_LIMIT = 20
IP_LOCK_MIN = 30


# -------------------------
# Time / hash
# -------------------------
def now_utc():
    return datetime.now(timezone.utc)

def hash_key(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


# -------------------------
# Text / parsing utilities
# -------------------------
def _norm_str(x):
    if x is None:
        return ""
    if isinstance(x, float) and pd.isna(x):
        return ""
    return str(x).strip()

def norm_key(x):
    # 공백/nbsp 제거 + strip (기관/회차 키 비교용)
    s = _norm_str(x).replace("\u00A0", " ")
    s = re.sub(r"\s+", "", s)
    return s.strip()

def extract_first_number_str(x):
    # "3학년", "고3", "3" -> "3"
    s = _norm_str(x)
    m = re.search(r"(\d+)", s)
    return m.group(1) if m else ""

def parse_percent_to_float(x):
    # '23.4', '23.4%', ' 23 %' 등 -> 23.4
    s = _norm_str(x)
    if s == "":
        return pd.NA
    m = re.search(r"(\d+(\.\d+)?)", s.replace(",", ""))
    if not m:
        return pd.NA
    try:
        return float(m.group(1))
    except Exception:
        return pd.NA

def to_int64_series(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce").astype("Int64")

def to_float(x):
    # "3.2", "3등급", "85%" -> float
    s = _norm_str(x)
    if s == "":
        return pd.NA
    m = re.search(r"(\d+(\.\d+)?)", s.replace(",", ""))
    if not m:
        return pd.NA
    try:
        return float(m.group(1))
    except Exception:
        return pd.NA


# -------------------------
# Question type mapping
# -------------------------
def build_question_type_map():
    m = {}
    def set_range(a, b, label):
        for q in range(a, b + 1):
            m[q] = label

    set_range(1, 17, "듣기")
    m[18] = "목적"
    m[19] = "심경"
    m[20] = "주장"
    m[21] = "함축적 의미"
    m[22] = "요지"
    m[23] = "주제"
    m[24] = "제목"
    m[25] = "표"
    m[26] = "지문 내용(세부)"
    set_range(27, 28, "실용자료(세부)")
    m[29] = "문법"
    m[30] = "어휘"
    set_range(31, 34, "빈칸 추론")
    m[35] = "글의 흐름"
    set_range(36, 37, "글의 순서")
    set_range(38, 39, "문장 삽입")
    m[40] = "문단 요약"
    m[41] = "제목(복합)"
    m[42] = "어휘(복합)"
    m[43] = "글의 순서(복합)"
    m[44] = "지칭 추론(복합)"
    m[45] = "내용일치/불일치(복합)"
    return m

QTYPE = build_question_type_map()

MAJOR_MAP = {
    "듣기": "듣기",
    "목적": "추론(단문)",
    "심경": "추론(단문)",
    "주장": "추론(단문)",
    "함축적 의미": "추론(단문)",
    "요지": "중심내용",
    "주제": "중심내용",
    "제목": "중심내용",
    "표": "세부정보",
    "지문 내용(세부)": "세부정보",
    "실용자료(세부)": "세부정보",
    "문법": "문법·어휘",
    "어휘": "문법·어휘",
    "어휘(복합)": "문법·어휘",  # 복합 어휘도 문법·어휘로 묶고 싶으면 유지
    "빈칸 추론": "빈칸 추론",
    "글의 흐름": "간접쓰기",
    "글의 순서": "간접쓰기",
    "문장 삽입": "간접쓰기",
    "문단 요약": "요약",
    "제목(복합)": "복합지문",
    "글의 순서(복합)": "복합지문",
    "지칭 추론(복합)": "복합지문",
    "내용일치/불일치(복합)": "복합지문",
}

MAJOR_COUNTS = {
    "듣기": 17,
    "추론(단문)": 4,
    "중심내용": 3,
    "세부정보": 4,
    "문법·어휘": 2,
    "빈칸 추론": 4,
    "간접쓰기": 5,
    "요약": 1,
    "복합지문": 5,
}


# -------------------------
# Wrong list parsing
# -------------------------
def parse_wrong_list(val):
    """
    returns (status, wrong_list[int], invalid_list[int])
    status: "응시" | "미응시" | "미입력"
    """
    s = _norm_str(val).replace(" ", "")
    if s == "":
        return "미입력", [], []
    if s == "미응시":
        return "미응시", [], []
    if s in {"0", "없음"}:
        return "응시", [], []

    parts = s.split(",")
    wrong, invalid = [], []
    for p in parts:
        if not p:
            continue
        if re.fullmatch(r"\d+", p):
            q = int(p)
            if 1 <= q <= 45:
                wrong.append(q)
            else:
                invalid.append(q)
    return "응시", sorted(set(wrong)), sorted(set(invalid))

def compute_major_counts(wrong_list):
    counts = {k: 0 for k in MAJOR_COUNTS.keys()}
    for q in wrong_list:
        detail = QTYPE.get(q, "기타")
        major = MAJOR_MAP.get(detail, "기타")
        if major in counts:
            counts[major] += 1
    return counts


# -------------------------
# Login attempt DB
# -------------------------
def ensure_auth_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS login_attempts (
            key TEXT PRIMARY KEY,
            fail_count INTEGER NOT NULL,
            first_fail_ts TEXT,
            last_fail_ts TEXT,
            locked_until_ts TEXT
        )
    """)
    conn.commit()
    return conn

def get_attempt(conn, key: str):
    cur = conn.cursor()
    cur.execute("SELECT fail_count, locked_until_ts FROM login_attempts WHERE key=?", (key,))
    row = cur.fetchone()
    if not row:
        return 0, None
    fail_count, locked_until_ts = row
    locked_until = None
    if locked_until_ts:
        try:
            locked_until = datetime.fromisoformat(locked_until_ts)
        except Exception:
            locked_until = None
    return fail_count, locked_until

def is_locked(conn, key: str):
    _, locked_until = get_attempt(conn, key)
    if locked_until and locked_until > now_utc():
        sec = int((locked_until - now_utc()).total_seconds())
        return True, sec
    return False, 0

def record_fail(conn, key: str, limit: int, lock_minutes: int):
    cur = conn.cursor()
    t = now_utc().isoformat()
    cur.execute("SELECT fail_count, locked_until_ts FROM login_attempts WHERE key=?", (key,))
    row = cur.fetchone()

    if not row:
        fail_count = 1
        locked_until = None
        if fail_count >= limit:
            locked_until = (now_utc() + timedelta(minutes=lock_minutes)).isoformat()
        cur.execute("""
            INSERT INTO login_attempts(key, fail_count, first_fail_ts, last_fail_ts, locked_until_ts)
            VALUES(?, ?, ?, ?, ?)
        """, (key, fail_count, t, t, locked_until))
    else:
        fail_count = int(row[0]) + 1
        locked_until_existing = row[1]
        locked_until_new = None
        if (not locked_until_existing) and fail_count >= limit:
            locked_until_new = (now_utc() + timedelta(minutes=lock_minutes)).isoformat()
        cur.execute("""
            UPDATE login_attempts
            SET fail_count=?, last_fail_ts=?, locked_until_ts=COALESCE(locked_until_ts, ?)
            WHERE key=?
        """, (fail_count, t, locked_until_new, key))

    conn.commit()

def reset_attempt(conn, key: str):
    cur = conn.cursor()
    cur.execute("DELETE FROM login_attempts WHERE key=?", (key,))
    conn.commit()

def get_client_ip_best_effort():
    try:
        headers = st.context.headers
        xff = headers.get("X-Forwarded-For")
        if xff:
            return xff.split(",")[0].strip()
        xrip = headers.get("X-Real-Ip")
        if xrip:
            return xrip.strip()
    except Exception:
        pass
    return "unknown"


# -------------------------
# Column standardizer
# -------------------------
def standardize_ebsi_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["출제기관", "회차", "학년", "문항번호", "전국오답률"])

    df = df.copy()
    colmap = {}
    for c in df.columns:
        cc = str(c).strip().replace("\u00A0", " ")
        cc_n = re.sub(r"\s+", "", cc)

        if cc_n in {"출제기관", "기관"}:
            colmap[c] = "출제기관"
        elif cc_n in {"회차", "시기", "시험", "모의고사"}:
            colmap[c] = "회차"
        elif cc_n in {"학년", "대상학년"}:
            colmap[c] = "학년"
        elif cc_n in {"문항번호", "문항", "번호"}:
            colmap[c] = "문항번호"
        elif cc_n in {"전국오답률", "전국오답률(%)", "오답률", "오답률(%)"}:
            colmap[c] = "전국오답률"

    df = df.rename(columns=colmap)

    for need in ["출제기관", "회차", "학년", "문항번호", "전국오답률"]:
        if need not in df.columns:
            df[need] = ""

    return df[["출제기관", "회차", "학년", "문항번호", "전국오답률"]].copy()

def standardize_grammar_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["출제기관", "회차", "학년", "문항번호", "정답개념"])

    df = df.copy()
    colmap = {}
    for c in df.columns:
        cc = str(c).strip().replace("\u00A0", " ")
        cc_n = re.sub(r"\s+", "", cc)

        if cc_n in {"출제기관", "기관"}:
            colmap[c] = "출제기관"
        elif cc_n in {"회차", "시기", "시험", "모의고사"}:
            colmap[c] = "회차"
        elif cc_n in {"학년", "대상학년"}:
            colmap[c] = "학년"
        elif cc_n in {"문항번호", "문항", "번호"}:
            colmap[c] = "문항번호"
        elif cc_n in {"정답개념", "개념", "정답개념명"}:
            colmap[c] = "정답개념"

    df = df.rename(columns=colmap)

    for need in ["출제기관", "회차", "학년", "문항번호", "정답개념"]:
        if need not in df.columns:
            df[need] = ""

    return df[["출제기관", "회차", "학년", "문항번호", "정답개념"]].copy()


# -------------------------
# Load Excel
# -------------------------
@st.cache_data(show_spinner=False)
def load_data(excel_path: str):
    students = pd.read_excel(excel_path, sheet_name="students")
    wrong = pd.read_excel(excel_path, sheet_name="wrong_answer")

    try:
        ebsi_raw = pd.read_excel(excel_path, sheet_name="ebsi_stats")
    except Exception:
        ebsi_raw = pd.DataFrame()

    try:
        grammar_raw = pd.read_excel(excel_path, sheet_name="grammar_info")
    except Exception:
        grammar_raw = pd.DataFrame()

    # ---- 3grade ----
    try:
        grade3 = pd.read_excel(excel_path, sheet_name="3grade", header=1)
    except Exception:
        grade3 = pd.DataFrame()

    # 컬럼명 정리(공백/nbsp 제거)
    grade3 = grade3.copy()
    grade3.columns = [
        str(c).replace("\u00A0", " ").strip()
        for c in grade3.columns
    ]

    # 혹시 "Unnamed: 0" 같은 찌꺼기 컬럼 제거
    grade3 = grade3.loc[:, ~grade3.columns.astype(str).str.match(r"^Unnamed")]

    # 필요한 컬럼이 없으면(시트가 비었거나 헤더가 꼬였거나) 빈 DF로 통일
    need_3 = ["이름", "모의고사응시횟수", "듣기영역(1~17번)", "독해영역(18~45번)", "등급평균"]
    if not set(need_3).issubset(set(grade3.columns)):
        # 여기서 바로 죽이지 말고, 학생 화면이 돌아가게만(로컬이니까)
        grade3 = pd.DataFrame(columns=need_3)

    # 값 클린(문자열)
    grade3["이름"] = grade3["이름"].astype(str).apply(lambda x: re.sub(r"\s+", "", x.replace("\u00A0", "")))
    
    # 숫자형
    grade3["모의고사응시횟수_num"] = grade3["모의고사응시횟수"].apply(to_float)
    grade3["등급평균_num"] = grade3["등급평균"].apply(to_float)




    try:
        admin_sol_raw = pd.read_excel(excel_path, sheet_name="admin_solution")
    except Exception:
        admin_sol_raw = pd.DataFrame()

    # ---- validate base sheets ----
    required_students = {"student_id", "name", "grade", "role"}
    if not required_students.issubset(set(students.columns)):
        raise ValueError(f"students 시트 컬럼 필요: {sorted(required_students)} / 현재: {list(students.columns)}")

    required_wrong = {"응시순서", "출제기관", "회차", "응시자", "원점수", "등급", "오답"}
    if not required_wrong.issubset(set(wrong.columns)):
        raise ValueError(f"wrong_answer 시트 컬럼 필요: {sorted(required_wrong)} / 현재: {list(wrong.columns)}")

    # ---- students clean ----
    students = students.copy()
    students["name"] = students["name"].astype(str).str.strip()
    students["student_id"] = students["student_id"].astype(str).str.strip()
    students["role"] = students["role"].astype(str).str.strip()
    students["grade"] = students["grade"].astype(str).str.strip()

    # ---- wrong_answer clean ----
    wrong = wrong.copy()
    wrong["응시자"] = wrong["응시자"].astype(str).str.strip()
    wrong["출제기관"] = wrong["출제기관"].astype(str).str.strip()
    wrong["회차"] = wrong["회차"].astype(str).str.strip()

    statuses, wrong_lists, invalid_lists = [], [], []
    for _, r in wrong.iterrows():
        stt, wl, inv = parse_wrong_list(r["오답"])

        raw = _norm_str(r["원점수"])
        grd = _norm_str(r["등급"])

        # 오답 미입력이지만 점수/등급이 있으면 응시로 처리
        if stt == "미입력":
            stt = "미입력" if (raw == "" and grd == "") else "응시"

        if _norm_str(r["원점수"]).replace(" ", "") == "미응시" or _norm_str(r["등급"]).replace(" ", "") == "미응시":
            stt, wl, inv = "미응시", [], []

        statuses.append(stt)
        wrong_lists.append(wl)
        invalid_lists.append(inv)

    wrong["status"] = statuses
    wrong["wrong_list"] = wrong_lists
    wrong["invalid_wrong_list"] = invalid_lists
    wrong["wrong_count"] = wrong["wrong_list"].apply(len)

    wrong["원점수_num"] = pd.to_numeric(wrong["원점수"].astype(str).str.extract(r"(\d+)")[0], errors="coerce")
    wrong["등급_num"] = pd.to_numeric(wrong["등급"].astype(str).str.extract(r"(\d+)")[0], errors="coerce")
    wrong["응시순서_num"] = pd.to_numeric(wrong["응시순서"], errors="coerce")

    df = wrong.merge(
        students[["student_id", "name", "grade", "role"]],
        left_on="응시자",
        right_on="name",
        how="left",
    )

    majors_df = pd.DataFrame(df["wrong_list"].apply(compute_major_counts).tolist())
    df = pd.concat([df.reset_index(drop=True), majors_df.reset_index(drop=True)], axis=1)

    # ---- ebsi normalize ----
    ebsi = standardize_ebsi_columns(ebsi_raw)
    ebsi = ebsi.copy()
    ebsi["출제기관_key"] = ebsi["출제기관"].apply(norm_key)
    ebsi["회차_key"] = ebsi["회차"].apply(norm_key)
    ebsi["학년_key"] = ebsi["학년"].apply(extract_first_number_str)
    ebsi["문항번호_num"] = to_int64_series(ebsi["문항번호"])
    ebsi["전국오답률_num"] = ebsi["전국오답률"].apply(parse_percent_to_float)

    # ---- grammar normalize ----
    grammar = standardize_grammar_columns(grammar_raw)
    grammar = grammar.copy()
    grammar["출제기관_key"] = grammar["출제기관"].apply(norm_key)
    grammar["회차_key"] = grammar["회차"].apply(norm_key)
    grammar["학년_key"] = grammar["학년"].apply(extract_first_number_str)
    grammar["문항번호_num"] = to_int64_series(grammar["문항번호"])
    grammar["정답개념"] = grammar["정답개념"].astype(str).str.replace("\u00A0", " ", regex=False).str.strip()
    grammar["정답개념_카테고리"] = grammar["정답개념"].astype(str).str.split("(", n=1).str[0].str.strip()

    # ---- admin_solution ----
    try:
        admin_sol = pd.read_excel(excel_path, sheet_name="admin_solution")
    except Exception:
        admin_sol = pd.DataFrame(columns=["name", "solution"])

    admin_sol = admin_sol.copy()
    admin_sol.columns = [str(c).replace("\u00A0", " ").strip() for c in admin_sol.columns]
    admin_sol = admin_sol.loc[:, ~admin_sol.columns.astype(str).str.match(r"^Unnamed")]

    need_admin = {"name", "solution"}
    if not need_admin.issubset(set(admin_sol.columns)):
        admin_sol = pd.DataFrame(columns=["name", "solution"])

    admin_sol["name"] = admin_sol["name"].astype(str).str.replace("\u00A0", " ", regex=False).str.strip()
    admin_sol["solution"] = admin_sol["solution"].astype(str).fillna("").str.replace("\u00A0", " ", regex=False).str.strip()


    return students, df, ebsi, grammar, grade3, admin_sol


# -------------------------
# Excel write (admin)
# -------------------------
def append_wrong_answer_row(excel_path: str, row: dict):
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl이 설치되어야 엑셀 저장이 가능합니다. 관리자에게 문의하세요.")

    wb = load_workbook(excel_path)
    if "wrong_answer" not in wb.sheetnames:
        raise ValueError("엑셀에 wrong_answer 시트가 없습니다.")
    ws = wb["wrong_answer"]

    headers = [cell.value for cell in ws[1]]
    for k in ["응시순서", "출제기관", "회차", "응시자", "원점수", "등급", "오답"]:
        if k not in headers:
            raise ValueError(f"wrong_answer 시트에 '{k}' 컬럼이 없습니다. 현재 헤더: {headers}")

    new_row = [row.get(h, "") for h in headers]
    ws.append(new_row)
    wb.save(excel_path)

def upsert_admin_solution(excel_path: str, name: str, solution: str):
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl이 설치되어야 엑셀 저장이 가능합니다. 관리자에게 문의하세요.")

    wb = load_workbook(excel_path)
    if "admin_solution" not in wb.sheetnames:
        ws = wb.create_sheet("admin_solution")
        ws.append(["name", "solution"])
    else:
        ws = wb["admin_solution"]

    headers = [cell.value for cell in ws[1]]
    # name/solution 컬럼 찾기(없으면 생성)
    if "name" not in headers and "solution" not in headers:
        # 기존 시트 헤더가 엉망이면 강제 표준화
        ws.delete_rows(1, 1)
        ws.insert_rows(1)
        ws.append(["name", "solution"])
        headers = ["name", "solution"]

    if "name" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="name")
        headers.append("name")
    if "solution" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="solution")
        headers.append("solution")

    name_col = headers.index("name") + 1
    sol_col = headers.index("solution") + 1

    name_norm = str(name).strip()
    sol_norm = str(solution).strip()

    # 기존 행 찾기
    found_row = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=name_col).value
        if v is None:
            continue
        if str(v).strip() == name_norm:
            found_row = r
            break

    if found_row is None:
        ws.append([""] * (len(headers)))
        found_row = ws.max_row
        ws.cell(row=found_row, column=name_col, value=name_norm)

    ws.cell(row=found_row, column=sol_col, value=sol_norm)
    wb.save(excel_path)


# -------------------------
# Login screen
# -------------------------
def render_login(students: pd.DataFrame):
    st.markdown("<h1 style='text-align:center;'>이루리 영어학원 방학 모의고사 성적분석</h1>", unsafe_allow_html=True)
    st.write("")

    name = st.text_input("이름")
    sid = st.text_input("고유번호", type="password")

    conn = ensure_auth_db()

    st.write("")
    if st.button("들어가기", type="primary"):
        if not name or not sid:
            st.error("이름과 고유번호를 모두 입력하세요.")
            return

        name_norm = str(name).strip()
        sid_norm = str(sid).strip()

        pair_key = hash_key(f"pair::{name_norm}::{sid_norm}")
        ip = get_client_ip_best_effort()
        ip_key = hash_key(f"ip::{ip}")

        locked, sec = is_locked(conn, pair_key)
        if locked:
            st.error(f"로그인 시도가 너무 많아서 잠겼습니다. {sec//60}분 {sec%60}초 후에 다시 시도하세요.")
            return

        locked, sec = is_locked(conn, ip_key)
        if locked and ip != "unknown":
            st.error(f"접속 시도가 너무 많아서 잠겼습니다. {sec//60}분 {sec%60}초 후에 다시 시도하세요.")
            return

        matched = students[
            (students["name"].astype(str).str.strip() == name_norm) &
            (students["student_id"].astype(str).str.strip() == sid_norm)
        ]

        if matched.empty:
            record_fail(conn, pair_key, PAIR_FAIL_LIMIT, PAIR_LOCK_MIN)
            if ip != "unknown":
                record_fail(conn, ip_key, IP_FAIL_LIMIT, IP_LOCK_MIN)
            st.error("이름 또는 고유번호를 확인해주세요.")
            return

        reset_attempt(conn, pair_key)
        if ip != "unknown":
            reset_attempt(conn, ip_key)

        role = str(matched.iloc[0]["role"]).strip()
        grade = matched.iloc[0].get("grade", None)

        st.session_state["logged_in"] = True
        st.session_state["role"] = role
        st.session_state["name"] = name_norm
        st.session_state["student_id"] = sid_norm
        st.session_state["grade"] = grade

        # 학생 화면 기본 상태
        st.session_state["student_panel"] = "none"  # none | total | exam
        st.session_state["admin_mode"] = "관리자 대시보드"

        st.rerun()


# -------------------------
# Student dashboard
# -------------------------
def render_student_dashboard(
    df: pd.DataFrame,
    ebsi: pd.DataFrame,
    grammar: pd.DataFrame,
    grade3: pd.DataFrame,
    admin_sol: pd.DataFrame,
    name: str,
    grade,
    is_preview: bool = False
):
    def build_wrong_table(wl, org, rnd, grd):
        tbl = pd.DataFrame({"문항번호": wl})
        if tbl.empty:
            return tbl

        tbl["문항번호"] = to_int64_series(tbl["문항번호"])
        tbl["유형"] = tbl["문항번호"].map(lambda q: QTYPE.get(int(q), "기타") if pd.notna(q) else "기타")
        tbl["대분류"] = tbl["유형"].map(lambda t: MAJOR_MAP.get(t, "기타"))

        org_k = norm_key(org)
        rnd_k = norm_key(rnd)
        g_k = extract_first_number_str(grd)

        # ---- EBSI: 학년 우선, 없으면 fallback ----
        base = ebsi[(ebsi["출제기관_key"] == org_k) & (ebsi["회차_key"] == rnd_k)].copy()
        use = base
        if g_k != "":
            base_g = base[base["학년_key"] == g_k]
            if not base_g.empty:
                use = base_g

        if not use.empty:
            rate_sub = use[["문항번호_num", "전국오답률_num"]].rename(columns={"문항번호_num": "문항번호"})
            tbl = tbl.merge(rate_sub, on="문항번호", how="left")
            tbl["전국오답률(%)"] = tbl["전국오답률_num"].apply(lambda x: "-" if pd.isna(x) else round(float(x), 1))
            tbl.drop(columns=["전국오답률_num"], inplace=True, errors="ignore")
        else:
            tbl["전국오답률(%)"] = "-"

        # ---- Grammar: 학년 우선, 없으면 fallback ----
        gbase = grammar[(grammar["출제기관_key"] == org_k) & (grammar["회차_key"] == rnd_k)].copy()
        guse = gbase
        if g_k != "":
            gbase_g = gbase[gbase["학년_key"] == g_k]
            if not gbase_g.empty:
                guse = gbase_g

        if not guse.empty:
            gsub = guse[["문항번호_num", "정답개념_카테고리", "정답개념"]].rename(columns={"문항번호_num": "문항번호"})
            tbl = tbl.merge(gsub, on="문항번호", how="left")
            tbl["정답개념_카테고리"] = tbl["정답개념_카테고리"].fillna("-")
            tbl["정답개념"] = tbl["정답개념"].fillna("-")
        else:
            tbl["정답개념_카테고리"] = "-"
            tbl["정답개념"] = "-"

        desired = ["문항번호", "유형", "대분류", "전국오답률(%)", "정답개념_카테고리", "정답개념"]
        for c in desired:
            if c not in tbl.columns:
                tbl[c] = "-"
        return tbl.loc[:, desired].copy()

    # header
    c1, c2 = st.columns([3, 1])
    with c1:
        suffix = " (미리보기)" if is_preview else ""
        st.markdown(f"### {name}{suffix}")
    with c2:
        st.markdown(
            f"<div style='text-align:right; font-size:16px; margin-top:8px;'>학년: <b>{grade if grade is not None else ''}</b></div>",
            unsafe_allow_html=True
        )

    me = df[df["응시자"] == name].copy()
    me = me.sort_values(["응시순서_num", "응시순서"], na_position="last") if "응시순서_num" in me.columns else me.sort_values(["응시순서"])
    taken = me[me["status"] == "응시"].copy()

    invalid_all = sorted(set([q for inv in taken["invalid_wrong_list"].tolist() for q in inv]))
    if invalid_all:
        st.warning(f"⚠️ 오답에 1~45 범위를 벗어난 번호가 있습니다: {invalid_all} (분석에서 제외)")

    # ----회차별 등급 추이 그래프 ----
    chart_df = taken.dropna(subset=["등급_num", "응시순서_num"]).copy()

    st.markdown("#### 회차별 등급 추이")
    if chart_df.empty:
        st.info("그래프를 그릴 데이터가 없습니다. (등급/응시순서 숫자 인식 실패 또는 응시 기록 없음)")
    else:
        fig = px.line(
            chart_df,
            x="응시순서_num",
            y="등급_num",
            markers=True,
            hover_data=["출제기관", "회차", "원점수_num", "wrong_count"],
        )
        fig.update_layout(showlegend=False)
        fig.update_xaxes(dtick=1, title="회차(응시순서)")
        fig.update_yaxes(autorange="reversed", dtick=1, range=[6.5, 0.5], title="등급(1~6)")
        st.plotly_chart(fig, use_container_width=True)

# ---- (수정된 KPI 로직: 공백 무시 매칭 버전) ----
    
    # 1. 비교를 위해 로그인한 이름에서 모든 공백 제거
    import re
    search_name = re.sub(r"\s+", "", str(name))
    
    # 2. grade3 시트에서 모든 공백을 제거한 이름과 비교하여 학생 찾기
    g3 = grade3[grade3["이름"].astype(str).apply(lambda x: re.sub(r"\s+", "", x)) == search_name].head(1)

    # 3. 기본값 설정
    total_attempts = "-"
    grade_avg = "-"
    listening_val = "-"
    reading_val = "-"

    if not g3.empty:
        r = g3.iloc[0]
        
        # 숫자 데이터 처리
        try:
            total_attempts = int(to_float(r.get("모의고사응시횟수", 0)))
            grade_avg = round(float(to_float(r.get("등급평균", 0))), 2)
        except:
            pass

        # 컬럼명을 유연하게 찾음 (듣기/독해 키워드 포함 컬럼)
        col_list = g3.columns.tolist()
        
        l_cols = [c for c in col_list if "듣기" in c]
        if l_cols:
            val = r[l_cols[0]]
            listening_val = str(val).strip() if pd.notna(val) and str(val).strip() != "" else "-"

        r_cols = [c for c in col_list if "독해" in c]
        if r_cols:
            val = r[r_cols[0]]
            reading_val = str(val).strip() if pd.notna(val) and str(val).strip() != "" else "-"
    else:
        # 데이터가 매칭되지 않았을 때 디버깅용 메시지 (나중에 삭제 가능)
        listening_val = "매칭 실패"
        reading_val = "매칭 실패"

    # 4. 화면 출력 (Metric)
    # ---- (1. 상단 4개 KPI 출력: 이전 답변의 커스텀 카드와 연결됨) ----
    st.markdown("""
        <style>
        .kpi-container {
            background-color: #ffffff;
            border-radius: 12px;
            padding: 20px;
            text-align: center;
            border: 2px solid #f0f2f6;
            box-shadow: 2px 4px 12px rgba(0,0,0,0.05);
            height: 100%;
            display: flex;
            flex-direction: column;
            justify-content: center;
        }
        .kpi-label { font-size: 14px; color: #555; margin-bottom: 10px; font-weight: 600; }
        .kpi-value { font-size: 18px; font-weight: 800; color: #1f77b4; word-break: break-all; }
        
        /* 화살표 스타일 */
        .flow-arrow {
            text-align: center;
            font-size: 35px;
            color: #1f77b4;
            margin: 20px 0;
            font-weight: bold;
            line-height: 1;
        }
        
        /* 솔루션 박스 강조 */
        .solution-box {
            border: 2px solid #1f77b4;
            border-radius: 15px;
            padding: 20px;
            background-color: #f0f8ff;
            box-shadow: 0 4px 15px rgba(31, 119, 180, 0.1);
        }
        </style>
    """, unsafe_allow_html=True)

    # 4개 박스 출력
    k1, k2, k3, k4 = st.columns(4)
    with k1: st.markdown(f'<div class="kpi-container"><div class="kpi-label">모의고사 응시 횟수</div><div class="kpi-value">{total_attempts}회</div></div>', unsafe_allow_html=True)
    with k2: st.markdown(f'<div class="kpi-container"><div class="kpi-label">듣기영역(1~17번)</div><div class="kpi-value">{listening_val}</div></div>', unsafe_allow_html=True)
    with k3: st.markdown(f'<div class="kpi-container"><div class="kpi-label">독해영역(18~45번)</div><div class="kpi-value">{reading_val}</div></div>', unsafe_allow_html=True)
    with k4: st.markdown(f'<div class="kpi-container"><div class="kpi-label">등급 평균</div><div class="kpi-value">{grade_avg}</div></div>', unsafe_allow_html=True)

    # ---- (2. 개선된 화살표 영역) ----
    st.markdown('<div class="flow-arrow">▼</div>', unsafe_allow_html=True)

    # ---- (3. 솔루션 박스 출력) ----
    sol_row = admin_sol[admin_sol["name"].astype(str).str.strip() == str(name).strip()].head(1)
    sol_text = _norm_str(sol_row.iloc[0].get("solution", "")) if not sol_row.empty else ""

    st.markdown("#### 🟦 솔루션")
    if sol_text.strip() == "":
        st.markdown('<div class="solution-box" style="color:#999;">작성된 솔루션이 없습니다.</div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div class="solution-box">{sol_text}</div>', unsafe_allow_html=True)

    st.divider()

    # ---- (위치변경) 버튼: 총오답 / 특정회차 ----
    b1, b2 = st.columns(2)
    with b1:
        if st.button("총 오답 현황 확인하기", key=f"btn_total_{name}"):
            st.session_state["student_panel"] = "total" if st.session_state.get("student_panel") != "total" else "none"
    with b2:
        if st.button("특정 회차 오답 확인하기", key=f"btn_exam_{name}"):
            st.session_state["student_panel"] = "exam" if st.session_state.get("student_panel") != "exam" else "none"

    panel = st.session_state.get("student_panel", "none")

    # ---- 총 오답 현황 ----
    if panel == "total":
        st.markdown("### 총 오답 현황 (회차별)")
        show = me[["응시순서", "출제기관", "회차", "status", "원점수_num", "등급_num", "wrong_count", "오답"]].copy()
        st.dataframe(show, use_container_width=True)

    # ---- 특정 회차 오답 ----
    if panel == "exam":
        st.markdown("### 특정 회차 오답")
        exams = taken[["응시순서_num", "응시순서", "출제기관", "회차"]].drop_duplicates().sort_values(["응시순서_num", "응시순서"])
        if not exams.empty:
            labels = [f"{int(r['응시순서'])}. {r['출제기관']} / {r['회차']}" for _, r in exams.iterrows()]
            choice = st.selectbox("회차", labels, key=f"exam_select_{name}")
            row = exams.iloc[labels.index(choice)]

            one = taken[taken["응시순서_num"] == row["응시순서_num"]].head(1)
            if not one.empty:
                wl = one.iloc[0]["wrong_list"]
                org = one.iloc[0]["출제기관"]
                rnd = one.iloc[0]["회차"]

                if not wl:
                    st.write("**틀린 문항:** 없음 (오답 0개)")
                else:
                    tbl = build_wrong_table(wl, org, rnd, grade)
                    st.dataframe(tbl, use_container_width=True)
                    st.caption("※ 전국 오답률: ebsi_stats / 문법 개념: grammar_info (학년 우선, 없으면 전체 fallback)")
        else:
            st.caption("응시 기록이 없습니다.")
    

    # ---- (버튼 삭제 후 항상 노출) 취약 유형 대분류 그래프 ----
    st.divider()
    st.markdown("### 취약 유형 (누적 대분류)")

    if taken.empty:
        st.info("응시 데이터가 없어서 분석할 수 없습니다.")
        return

    major_cols = list(MAJOR_COUNTS.keys())
    sums = taken[major_cols].sum().sort_values(ascending=False)
    st.bar_chart(sums)

    # ---- (추가) 문법·어휘 오답 10개 초과 시 '문법 개념 키워드' ----
    gram_vocab_wrong = int(sums.get("문법·어휘", 0))

    if gram_vocab_wrong > 10:
        st.markdown("#### ⚠️ 문법·어휘 오답이 많습니다 (10개 초과)")
        # 누적 오답에서 문법 관련 문항만 뽑아 개념 키워드 집계
        all_wrong = []
        for wl in taken["wrong_list"]:
            all_wrong.extend(wl)

        # 문법/어휘 관련 번호만(29, 30, 42 정도)
        gv_set = set([29, 30, 42])
        gv_wrong = [q for q in all_wrong if q in gv_set]

        if not gv_wrong:
            st.caption("문법·어휘로 분류된 오답은 있으나, 문항번호(29/30/42)에서 직접 확인되지 않았습니다.")
        else:
            # 가장 최근 회차 기준으로 개념 매핑을 얻는 쪽이 자연스럽지만,
            # 여기서는 누적 키워드 제시가 목적이라 '현재 존재하는 매핑'에서 최대한 모음
            # (출제기관/회차별로 개념이 다를 수 있음)
            # 간단히: grammar_info에서 해당 문항번호의 카테고리 빈도 집계
            cats = []
            for _, row in taken.iterrows():
                org_k = norm_key(row["출제기관"])
                rnd_k = norm_key(row["회차"])
                g_k = extract_first_number_str(grade)

                gbase = grammar[(grammar["출제기관_key"] == org_k) & (grammar["회차_key"] == rnd_k)].copy()
                guse = gbase
                if g_k != "":
                    gbase_g = gbase[gbase["학년_key"] == g_k]
                    if not gbase_g.empty:
                        guse = gbase_g

                if guse.empty:
                    continue

                wrongs = set(row["wrong_list"])
                sub = guse[guse["문항번호_num"].isin(list(wrongs))][["정답개념_카테고리"]].copy()
                for v in sub["정답개념_카테고리"].dropna().astype(str).tolist():
                    vv = v.strip()
                    if vv and vv != "-":
                        cats.append(vv)

            if not cats:
                st.caption("grammar_info에 매핑된 '정답개념' 데이터가 부족해서 키워드를 만들 수 없습니다.")
                st.caption("→ grammar_info 시트에 (출제기관/회차/학년/문항번호/정답개념) 채우면 자동으로 뜹니다.")
            else:
                top = pd.Series(cats).value_counts().head(8)
                st.write("**틀린 문법 개념 키워드(상위):**")
                st.write(" · ".join([f"{idx}({int(val)})" for idx, val in top.items()]))

    # (요구사항) 최근 성적요약 / 기존 솔루션 / 성적분석 탭은 삭제됨


# -------------------------
# Admin dashboard
# -------------------------
def render_admin_dashboard(df: pd.DataFrame, students_df: pd.DataFrame, admin_sol: pd.DataFrame):
    st.markdown("### 관리자 모드")
    st.caption("학생 개별 조회 + 학원 전체 취약 유형(랭킹/비교 없음)")

    taken = df[df["status"] == "응시"].copy()
    absent = df[df["status"] == "미응시"].copy()
    missing = df[df["status"] == "미입력"].copy()

    k1, k2, k3 = st.columns(3)
    k1.metric("응시 기록(행)", len(taken))
    k2.metric("미응시(행)", len(absent))
    k3.metric("미입력/누락(행)", len(missing))

    invalid_all = sorted(set([q for inv in taken["invalid_wrong_list"].tolist() for q in inv]))
    if invalid_all:
        st.warning(f"⚠️ 오답에 1~45 범위를 벗어난 번호가 있습니다: {invalid_all} (분석에서 제외)")

    st.divider()

    st.markdown("#### 학원 전체 취약 유형 (대분류)")
    if taken.empty:
        st.info("응시 데이터가 없어서 집계할 수 없습니다.")
    else:
        major_cols = list(MAJOR_COUNTS.keys())
        sums = taken[major_cols].sum().sort_values(ascending=False)
        st.bar_chart(sums)

        st.markdown("#### 학원 내 대분류 오답률(%) (전체 분포)")
        n_attempts = len(taken)
        rates = {}
        for major, qcnt in MAJOR_COUNTS.items():
            denom = qcnt * n_attempts
            num = float(taken[major].sum())
            rates[major] = (num / denom * 100.0) if denom else 0.0

        rate_df = (
            pd.DataFrame({
                "대분류": list(rates.keys()),
                "오답률(%)": [round(v, 2) for v in rates.values()],
            })
            .sort_values("오답률(%)", ascending=False)
            .reset_index(drop=True)
        )
        st.dataframe(rate_df, use_container_width=True)

    # ---- (추가) 관리자 솔루션 작성 칸 ----
    st.divider()
    st.markdown("#### (관리자) 관리자 솔루션 작성 (학생 화면에 표시됨)")

    if not HAS_OPENPYXL:
        st.warning("엑셀 저장 기능을 사용하려면 openpyxl이 필요합니다.")
    else:
        students_list = sorted(students_df["name"].dropna().astype(str).str.strip().unique().tolist())
        target = st.selectbox("학생 선택", students_list, key="sol_student")

        current = admin_sol[admin_sol["name"].astype(str).str.strip() == str(target).strip()].head(1)
        current_text = "" if current.empty else _norm_str(current.iloc[0].get("solution", ""))

        text = st.text_area("관리자 솔루션", value=current_text, height=150, placeholder="예: 듣기에서 어조/태도 유형을 집중 보완합시다...")

        c1, c2 = st.columns([1, 3])
        with c1:
            if st.button("저장", type="primary", key="btn_save_sol"):
                try:
                    upsert_admin_solution(EXCEL_PATH, target, text)
                    st.cache_data.clear()
                    st.success("저장 완료! (학생 화면에 바로 반영됩니다)")
                    st.rerun()
                except PermissionError:
                    st.error("엑셀 파일이 열려있어서 저장할 수 없습니다. 엑셀/미리보기 닫고 다시 시도하세요.")
                except Exception as e:
                    st.error(f"저장 실패: {e}")
        with c2:
            st.caption("※ 로컬 전용. DB아님.")

    st.divider()

    st.markdown("#### (관리자) 응시 기록 추가")
    if not HAS_OPENPYXL:
        st.warning("엑셀 저장 기능을 사용하려면 openpyxl이 필요합니다. 개발자에게 문의하세요.")
    else:
        students_list = sorted(students_df["name"].dropna().astype(str).str.strip().unique().tolist())

        with st.form("add_exam_form", clear_on_submit=True):
            col1, col2, col3 = st.columns(3)
            with col1:
                new_order = st.text_input("응시순서(숫자)", placeholder="예: 5")
                ORG_OPTIONS = ["평가원", "교육청", "사설", "기타"]
                org_choice = st.selectbox("출제기관", ORG_OPTIONS, index=0)
                new_org = st.text_input("출제기관 직접 입력") if org_choice == "기타" else org_choice
            with col2:
                new_round = st.text_input("회차", placeholder="예: 25년 3월")
                new_name = st.selectbox("응시자(이름)", students_list, index=0)
            with col3:
                new_score = st.text_input("원점수", placeholder="예: 92")
                new_grade = st.text_input("등급", placeholder="예: 2 또는 2등급")

            new_wrong = st.text_input("오답(쉼표구분)", placeholder="예: 3,5,12,29 / 없으면 0 / 미응시는 '미응시'")
            submitted = st.form_submit_button("추가 저장", type="primary")

        if submitted:
            if not new_order.strip() or not re.fullmatch(r"\d+", new_order.strip()):
                st.error("응시순서는 숫자로 입력해야합니다. (예: 5)")
            elif not new_org.strip() or not new_round.strip() or not new_name.strip():
                st.error("출제기관/회차/응시자는 필수작성 항목입니다.")
            else:
                try:
                    row = {
                        "응시순서": new_order.strip(),
                        "출제기관": new_org.strip(),
                        "회차": new_round.strip(),
                        "응시자": new_name.strip(),
                        "원점수": new_score.strip(),
                        "등급": new_grade.strip(),
                        "오답": new_wrong.strip(),
                    }
                    append_wrong_answer_row(EXCEL_PATH, row)
                    st.cache_data.clear()
                    st.success("저장 완료! 새로고침합니다.")
                    st.rerun()
                except PermissionError:
                    st.error("엑셀 파일이 열려있어서 저장 할 수 없습니다. 엑셀/미리보기 닫고 다시 시도하십시오.")
                except Exception as e:
                    st.error(f"저장 실패: {e}")

    st.divider()

    st.markdown("#### 학생별 조회")
    students = sorted([n for n in df["응시자"].dropna().unique().tolist()])
    selected = st.selectbox("학생 선택", students, key="admin_student_select")
    if selected:
        sub = df[df["응시자"] == selected].sort_values(["응시순서_num", "응시순서"], na_position="last")
        st.dataframe(
            sub[["응시순서", "출제기관", "회차", "status", "원점수_num", "등급_num", "wrong_count", "오답"]],
            use_container_width=True
        )


# -------------------------
# Main
# -------------------------
def main():
    try:
        students, df, ebsi, grammar, grade3, admin_sol = load_data(EXCEL_PATH)
    except Exception as e:
        st.error(f"엑셀 로드 실패: {e}")
        st.stop()

    if "logged_in" not in st.session_state:
        st.session_state["logged_in"] = False

    if not st.session_state["logged_in"]:
        render_login(students)
        return

    role = str(st.session_state.get("role", "")).strip()
    name = st.session_state.get("name")
    grade = st.session_state.get("grade")

    with st.sidebar:
        st.markdown("## 계정")
        st.write(f"- 이름: **{st.session_state.get('name')}**")
        st.write(f"- 역할: **{st.session_state.get('role')}**")

        if role == "admin":
            st.markdown("## 관리자 메뉴")
            st.session_state["admin_mode"] = st.radio(
                "화면 선택",
                ["관리자 대시보드", "학생 화면 미리보기"],
                index=0 if st.session_state.get("admin_mode") != "학생 화면 미리보기" else 1
            )

        if st.button("로그아웃"):
            st.session_state.clear()
            st.rerun()

    if role == "admin":
        mode = st.session_state.get("admin_mode", "관리자 대시보드")
        if mode == "관리자 대시보드":
            render_admin_dashboard(df, students, admin_sol)
        else:
            st.markdown("### 학생 화면 미리보기(관리자)")
            students_list = sorted(students["name"].dropna().astype(str).str.strip().unique().tolist())
            preview_name = st.selectbox("미리볼 학생 선택", students_list, key="preview_student")

            g = students[students["name"].astype(str).str.strip() == str(preview_name).strip()]
            preview_grade = g.iloc[0]["grade"] if not g.empty else ""

            render_student_dashboard(df, ebsi, grammar, grade3, admin_sol, name=preview_name, grade=preview_grade, is_preview=True)
    else:
        render_student_dashboard(df, ebsi, grammar, grade3, admin_sol, name=name, grade=grade, is_preview=False)


if __name__ == "__main__":
    main()
